import os
import django
import openpyxl as xl
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'dmoj.settings')
django.setup()

from judge.models import Profile,Organization,Language
from judge.utils.matrix_utils import MatrixUtils
from django.contrib.auth.models import User

wb = xl.load_workbook("./misc/taikhoang.xlsx")
Matrix = MatrixUtils()

def init(org_name:str):
    org = Organization.objects.get_or_create(name=org_name.replace(' ','_'))
    if(org[1]):
        return
    org = org[0]
    org.is_open = False
    org.name = org_name
    org.slug = org_name.replace(' ','_')
    org.short_name = org_name
    org.about = org_name
    org.admins.add(Profile.objects.get(user=1))
    org.save()
    Matrix.create_room(org_name)
    return org

def create(username,password,email,fullname,org,**kwarg):
    Matrix.create_account(username,fullname,password)
    Matrix.add_user(username,org.name)
    obj = User(username=username,email=email,first_name=fullname,is_active=True)
    obj.set_password(password)
    obj.save()
    prof = Profile(user=obj)
    prof.language = Language.objects.get(key="cpp14")
    
    prof.save()
    prof.organizations.add(org)
    prof.save()

data = wb['Sheet1']
conf = wb['Sheet2']
startpoint = conf.cell(1,1).value
number_of_account = conf.cell(1,2).value
names = []

org = init(conf.cell(3,2).value)

for index in conf[2]:
    names.append(index.value.lower().replace(' ',''))
print(names)
for index in range(startpoint,startpoint + number_of_account):
    user_data = []
    for infomation in range(len(names)):
        user_data.append(data.cell(index,infomation+1).value)
    print(user_data)
    data_maps = {}
    for name_index in range(len(names)):
        data_maps[names[name_index]] = user_data[name_index]
    create(org=org,**data_maps)
    

print('------- done -------')